from openpyxl import Workbook, load_workbook
import os

# Helper untuk cek dan buat file Excel jika belum ada
def init_excel_files():
    if not os.path.exists("zakat_data.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Nama", "Jenis Zakat", "Jumlah", "Tanggal"])
        wb.save("zakat_data.xlsx")
    
    if not os.path.exists("master_beras.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Nama Beras", "Harga per Kg"])
        wb.save("master_beras.xlsx")
        
    if not os.path.exists("transaksi_zakat.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "ID Zakat", "ID Beras", "Jumlah Beras", "Total Harga", "Tanggal"])
        wb.save("transaksi_zakat.xlsx")

def get_next_id(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    return len(ws["A"])

def add_zakat(nama, jenis_zakat, jumlah, tanggal):
    file = "zakat_data.xlsx"
    wb = load_workbook(file)
    ws = wb.active
    id = get_next_id(file)
    ws.append([id, nama, jenis_zakat, jumlah, tanggal])
    wb.save(file)

def update_zakat(id, nama, jenis_zakat, jumlah, tanggal):
    wb = load_workbook("zakat_data.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == id:
            row[1].value = nama
            row[2].value = jenis_zakat
            row[3].value = jumlah
            row[4].value = tanggal
            break
    wb.save("zakat_data.xlsx")

def delete_zakat(id):
    wb = load_workbook("zakat_data.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == id:
            ws.delete_rows(row[0].row)
            break
    wb.save("zakat_data.xlsx")

def add_beras(nama_beras, harga_per_kg):
    file = "master_beras.xlsx"
    wb = load_workbook(file)
    ws = wb.active
    id = get_next_id(file)
    ws.append([id, nama_beras, harga_per_kg])
    wb.save(file)

def view_master_beras():
    wb = load_workbook("master_beras.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(f"ID: {row[0]}, Nama Beras: {row[1]}, Harga per Kg: {row[2]}")

def add_transaksi_zakat(id_zakat, id_beras, jumlah_beras, tanggal):
    # Ambil harga beras
    wb_beras = load_workbook("master_beras.xlsx")
    ws_beras = wb_beras.active
    harga_per_kg = None
    for row in ws_beras.iter_rows(min_row=2):
        if row[0].value == id_beras:
            harga_per_kg = row[2].value
            break
    if harga_per_kg is None:
        print("ID beras tidak ditemukan.")
        return
    total_harga = jumlah_beras * harga_per_kg
    
    file = "transaksi_zakat.xlsx"
    wb = load_workbook(file)
    ws = wb.active
    id_transaksi = get_next_id(file)
    ws.append([id_transaksi, id_zakat, id_beras, jumlah_beras, total_harga, tanggal])
    wb.save(file)
    print("Transaksi zakat berhasil ditambahkan.")

def view_transaksi_zakat():
    wb_trx = load_workbook("transaksi_zakat.xlsx")
    ws_trx = wb_trx.active
    wb_zakat = load_workbook("zakat_data.xlsx")
    ws_zakat = wb_zakat.active
    wb_beras = load_workbook("master_beras.xlsx")
    ws_beras = wb_beras.active

    # Buat dict bantu
    zakat_dict = {row[0].value: (row[1].value, row[2].value) for row in ws_zakat.iter_rows(min_row=2)}
    beras_dict = {row[0].value: row[1].value for row in ws_beras.iter_rows(min_row=2)}

    for row in ws_trx.iter_rows(min_row=2, values_only=True):
        id_trx, id_zakat, id_beras, jumlah, total, tanggal = row
        nama_zakat, jenis = zakat_dict.get(id_zakat, ("?", "?"))
        nama_beras = beras_dict.get(id_beras, "?")
        print(f"ID Transaksi: {id_trx}, Nama: {nama_zakat}, Jenis: {jenis}, Beras: {nama_beras}, "
              f"Jumlah: {jumlah} kg, Total: {total}, Tanggal: {tanggal}")

def input_master_beras():
    print("\nTambah Data Master Beras")
    nama_beras = input("Masukkan nama jenis beras: ")
    harga_per_kg = float(input("Masukkan harga per kg: "))
    add_beras(nama_beras, harga_per_kg)
    print("Data master beras berhasil ditambahkan!")

def main():
    init_excel_files()
    while True:
        print("\nMenu:")
        print("1. Tambah Data Zakat")
        print("2. Edit Data Zakat")
        print("3. Hapus Data Zakat")
        print("4. Lihat Data Master Beras")
        print("5. Tambah Data Master Beras")
        print("6. Tambah Transaksi Zakat")
        print("7. Lihat Transaksi Zakat")
        print("8. Keluar")

        choice = input("Pilih opsi (1-8): ")
        
        if choice == "1":
            nama = input("Masukkan nama: ")
            jenis_zakat = input("Masukkan jenis zakat: ")
            jumlah = float(input("Masukkan jumlah zakat: "))
            tanggal = input("Masukkan tanggal (YYYY-MM-DD): ")
            add_zakat(nama, jenis_zakat, jumlah, tanggal)
            print("Data zakat berhasil ditambahkan.")
        
        elif choice == "2":
            id_zakat = int(input("Masukkan ID zakat yang ingin diubah: "))
            nama = input("Masukkan nama baru: ")
            jenis_zakat = input("Masukkan jenis zakat baru: ")
            jumlah = float(input("Masukkan jumlah zakat baru: "))
            tanggal = input("Masukkan tanggal baru (YYYY-MM-DD): ")
            update_zakat(id_zakat, nama, jenis_zakat, jumlah, tanggal)
            print("Data zakat berhasil diperbarui.")
        
        elif choice == "3":
            id_zakat = int(input("Masukkan ID zakat yang ingin dihapus: "))
            delete_zakat(id_zakat)
            print("Data zakat berhasil dihapus.")
        
        elif choice == "4":
            print("\nMaster Data Beras:")
            view_master_beras()
        
        elif choice == "5":
            input_master_beras()
        
        elif choice == "6":
            id_zakat = int(input("Masukkan ID zakat: "))
            id_beras = int(input("Masukkan ID beras: "))
            jumlah_beras = float(input("Masukkan jumlah beras (kg): "))
            tanggal = input("Masukkan tanggal (YYYY-MM-DD): ")
            add_transaksi_zakat(id_zakat, id_beras, jumlah_beras, tanggal)
        
        elif choice == "7":
            print("\nTransaksi Zakat:")
            view_transaksi_zakat()
        
        elif choice == "8":
            print("Keluar dari program.")
            break
        
        else:
            print("Pilihan tidak valid. Silakan coba lagi.")

main()